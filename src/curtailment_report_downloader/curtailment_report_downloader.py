import io
import pycurl
from pathlib import Path
import pandas as pd
from pandas import Timestamp as ts
from datetime import date as d, timedelta as td
from openpyxl import load_workbook

from src.logging.logging import DataLogger,TextLogger

class CurtailmentReportDownloader:
    '''
    A class to manage downloads of CAISO daily curtailment reports.
    '''
    start_date = d(2021,6,18)
    column_names = [
        'OUTAGE MRID',
        'RESOURCE NAME',
        'RESOURCE ID',
        'OUTAGE TYPE',
        'NATURE OF WORK',
        'CURTAILMENT START DATE TIME',
        'CURTAILMENT END DATE TIME',
        'CURTAILMENT MW',
        'RESOURCE PMAX MW',
        'NET QUALIFYING CAPACITY MW',
        'OUTAGE STATUS',
        'RES TYPE',
        'MKTORGANIZATION MRID',
        'BAA'
    ]
    def __init__(self,config:dict):
        log_dtypes = {
            'effective_date' : 'datetime64[D]',
            'source_url' : 'string',
            'download_path' : 'string',
            'loaded_to_parquet' : 'int64',
        }

        log_path = Path(config['caiso_curtailment_reports']['download_log_path'])

        self.urls = config['caiso_curtailment_reports']['url']
        self.download_path_template = config['caiso_curtailment_reports']['download_path_template']
        self.combined_reports_path = Path(config['caiso_curtailment_reports']['combined_reports_path'])
        self.logger = DataLogger(dtypes=log_dtypes,log_path=log_path,delimiter=',')
        self.status_logger = TextLogger(
            cli_logging_criticalities=['INFORMATION','WARNING','ERROR'],
            file_logging_criticalities=['WARNING','ERROR'],
            log_path=config['caiso_curtailment_reports']['text_log_path']
        )

    def url_by_date(self,date:d):
        '''
        Generates the url for a CAISO prior trade day curtailment report for the
        input date, based on templates for a standard url and exceptions defined
        in the config file.

        Parameters:
            date - a Pandas Timestamp object representing a single day.

        Returns:
            A string url pointing to a report if it exists for the given day.
        '''
        
        # Retrieve the standard or exceptional url template from config file:
        for e in self.urls['exceptions']:
            if (e['type']=='list' and date in e['dates']) or \
                (e['type']=='range' and
                    date>=e['dates'][0] and date<=e['dates'][1]):
                date_template = e['template']
                break
        else:
            date_template = self.urls['standard']
        
        # Apply the input date for the selected template to return full url:
        return date.strftime(date_template)

    def download_path_by_date(self,effective_date:d):
        '''
        Generates the path to which a report is or will be saved for the given
        date.

        Parameters:
            date - a date object representing a given day

        Returns:
            A pathlib Path object pointing to a local location for a report to
            be saved.
        '''
        return Path(effective_date.strftime(self.download_path_template))

    def download_report_by_date(self,effective_date:d):
        '''
        Downloads a prior trade day curtailments report from the CAISO website,
        if available.

        Parameters:
            date - a date object representing a given day

        Side Effects:
            Downloads and saves an Excel spreadsheet file.
            Prints actions to console
            Appends the download log

        Returns:
            Integer representing status of download:
                -1 if unable to download file (response from CAISO website != 200)
                0 if file is already downloaded according to log
                200 if file successfully downloaded and saved to location
        '''
        if (ts(effective_date)==self.logger.data.loc[:,'effective_date']).any():
            download_date = self.logger.data.loc[(self.logger.data.loc[:,'effective_date']==ts(effective_date)),'log_timestamp'].iloc[0]
            self.status_logger.log('Skipping Download for {} [Already downloaded {}]'.format(effective_date.strftime('%Y-%m-%d'),download_date.strftime('%Y-%m-%d %H:%M:%S')))
            return 0
        else:
            url = self.url_by_date(effective_date)
            download_path = self.download_path_by_date(effective_date)
            with download_path.open('wb') as f:
                c = pycurl.Curl()
                c.setopt(c.URL,url)
                c.setopt(c.FOLLOWLOCATION,True)
                c.setopt(c.WRITEDATA,f)
                c.perform()
                response_code = c.getinfo(c.RESPONSE_CODE)
                c.close()
                self.logger.log(pd.Series({
                    'effective_date' : effective_date,
                    'source_url' : url,
                    'download_path' : download_path,
                    'loaded_to_parquet' : 0,
                }))
                self.logger.commit()
            self.status_logger.log('Downloading for {}: {}'.format(effective_date.strftime('%Y-%m-%d'),url))
            return response_code

    def download_all_reports(self):
        '''
        Downloads all prior trade day curtailments report from the CAISO website,
        from the initial date available (June 18, 2021) to yesterday.

        Parameters:
            None

        Side Effects:
            Calls download_report_by_date method.

        Returns:
            None
        '''
        today = d.today()
        date_range = [
            d.fromordinal(x) for x in \
            range(self.start_date.toordinal(),today.toordinal())
        ]
        download_count = 0
        skip_count = 0
        error_dates = []
        for date in date_range:
            result = self.download_report_by_date(date)
            if result==200:
                download_count += 1
            elif result<0:
                skip_count += 1
            else:
                error_dates.append(date)
        if download_count>1:
            self.status_logger.log(f'Downloaded {download_count} new reports.')
        elif download_count>0:
            self.status_logger.log(f'Downloaded {download_count} new report.')
        else:
            self.status_logger.log('No new reports downloaded.')
        if skip_count>1:
            self.status_logger.log(f'Skipped {skip_count} reports already downloaded.')
        elif skip_count>0:
            self.status_logger.log(f'Skipped {skip_count} report already downloaded.')
        else:
            self.status_logger.log('No reports skipped')
        if len(error_dates)>1:
            for error_date in error_dates:
                self.status_logger.log(
                    'Unable to download reports for the following dates: ' + \
                    ', '.join([error_date.strftime('%Y-%m-%d') for error_date in error_dates]),
                    criticality='WARNING'
                )
        elif len(error_dates)>0:
            self.status_logger.log(
                'Unable to download report for the following date: '
                + error_dates[0].strftime('%Y-%m-%d'),
                criticality='WARNING'
            )

    def load_parquet(self):
        '''
        Loads data from a saved parquet file to reduce time reading from excel
        files
        '''
        if self.combined_reports_path.is_file():
            return pd.read_parquet(self.combined_reports_path)
        else:
            return pd.DataFrame(columns=['REPORT DATE']+self.column_names)

    def dump_parquet(self):
        '''
        Saves the current dataframe of curtailment reports to a parquet file at
        the path specified in the config dictionary
        '''
        self.curtailment_data.to_parquet(self.combined_reports_path)
    
    def clear_parquet(self):
        '''
        Deletes the parquet file containing combined curtailment reports and
        sets the value of the 'loaded_to_parquet' column in the log to 0 for all
        downloaded reports
        '''
        self.logger.data.loc[:,'loaded_to_parquet'] = 0
        self.logger.commit()
        self.combined_reports_path.unlink()

    def extract_report_by_date(self,effective_date:d):
        '''
        Extracts data from a single report into a pandas dataframe corresponding
        to the input effective date, and downloads the report if it is not
        locally available.

        Parameters:
            effective_date - a singled datetime object 
        
        Returns:
            DataFrame containing contents of the selected prior trade-day
            curtailment report.
        '''
        if (ts(effective_date)!=self.logger.data.loc[:,'effective_date']).all():
            self.download_report_by_date(effective_date)

        df = pd.DataFrame(columns=self.column_names)
        download_path = Path(self.download_path_by_date(effective_date))
        
        # open downloaded report workbook:
        with download_path.open('rb') as f:
            self.status_logger.log('Reading ' + download_path.name)
            in_mem_file = io.BytesIO(f.read())
            wb = load_workbook(in_mem_file,data_only=True,read_only=True)
        ws = wb['PREV_DAY_OUTAGES']
        new_data = {k:[] for k in self.column_names}

        # find header row based on text matching:
        header_row_number = 1
        while True:
            header_row = list(map(lambda x:x.value,ws[header_row_number]))
            if self.column_names[0] in header_row or header_row_number>100:
                break
            else:
                header_row_number += 1

        columns = {k: header_row.index(k) if k in header_row else None for k in self.column_names}
        if header_row_number<100:
            for data_range_row in ws.iter_rows(min_row=header_row_number+1):
                if len(data_range_row)>0:
                    for column_name,column_number in columns.items():
                        if column_number is not None:
                            new_data[column_name].append(data_range_row[column_number].value)
                        else:
                            new_data[column_name].append(None)
            new_dataframe = pd.DataFrame(new_data)

            # Constrain curtailment hours within trade day:
            # new_dataframe.loc[:,'CURTAILMENT START DATE TIME'] = new_dataframe.loc[:,'CURTAILMENT START DATE TIME'].apply(lambda t: max(t,ts(effective_date)))
            # new_dataframe.loc[:,'CURTAILMENT END DATE TIME'] = new_dataframe.loc[:,'CURTAILMENT END DATE TIME'].fillna(effective_date+td(days=1))
            # new_dataframe.loc[:,'CURTAILMENT END DATE TIME'] = new_dataframe.loc[:,'CURTAILMENT END DATE TIME'].apply(lambda t: min(t,ts(effective_date)))

            # Append new data to dataframe:
            df = df.append(new_dataframe,ignore_index=True)
        return df
    
    def update_parquet(self):
        '''
        Loads data from the combined reports parquet file, then extracts any
        downloaded reports not already loaded and saves the updated DataFrame
        to the parquet file.
        '''
        df = self.load_parquet()
        unloaded_reports = self.logger.data.loc[(self.logger.data.loc[:,'loaded_to_parquet']==0),:]
        for _,r in unloaded_reports.iterrows():
            new_data = self.extract_report_by_date(r.loc['effective_date'].date())
            columns = new_data.columns
            new_data.loc[:,'REPORT DATE'] = r.loc['effective_date']
            new_data.loc[:,['REPORT DATE']+list(columns)]
            df = pd.concat([df,new_data],ignore_index=True)
            self.logger.data.loc[self.logger.data.loc[:,'effective_date']==r.loc['effective_date'],'loaded_to_parquet'] = 1
        df.to_parquet(self.combined_reports_path)
        self.logger.commit()

    def extract_all_reports(self,effective_dates:list=[]):
        '''
        Extracts data from all downloaded reports without filtering.

        Parameters:
            effective_dates - a list containing datetime objects representing
                dates corresponding to the effective_date column of the log
                associated with filenames from which to extract data

        Returns:
            Dataframe containing data from curtailment reports matching
            the given effective dates.
        '''
        df = pd.DataFrame(columns=self.column_names)
        if len(effective_dates)>0:
            download_path_strs = list(self.logger.data.loc[self.logger.data.loc[:,'effective_date'].apply(lambda t: t in effective_dates),'download_path'])
        else:
            download_path_strs = list(self.logger.data.loc[:,'download_path'])
        for effective_date,download_path_str in zip(effective_dates,download_path_strs):
            with Path(download_path_str).open('rb') as f:
                self.status_logger.log('Reading '+Path(download_path_str).name)
                in_mem_file = io.BytesIO(f.read())
                wb = load_workbook(in_mem_file,data_only=True,read_only=True)
            ws = wb['PREV_DAY_OUTAGES']
            new_data = {k:[] for k in self.column_names}
            # find header row:
            header_row_number = 1
            while True:
                header_row = list(map(lambda x:x.value,ws[header_row_number]))
                if self.column_names[0] in header_row or header_row_number>100:
                    break
                else:
                    header_row_number += 1

            columns = {k: header_row.index(k) if k in header_row else None for k in self.column_names}
            if header_row_number<100:
                for data_range_row in ws.iter_rows(min_row=header_row_number+1):
                    if len(data_range_row)>0:
                        for column_name,column_number in columns.items():
                            if column_number is not None:
                                new_data[column_name].append(data_range_row[column_number].value)
                            else:
                                new_data[column_name].append(None)
                new_dataframe = pd.DataFrame(new_data)
                # Constrain curtailment hours within trade day:
                new_dataframe.loc[:,'CURTAILMENT START DATE TIME'] = new_dataframe.loc[:,'CURTAILMENT START DATE TIME'].apply(lambda t: max(t,effective_date))
                new_dataframe.loc[:,'CURTAILMENT END DATE TIME'] = new_dataframe.loc[:,'CURTAILMENT END DATE TIME'].fillna(effective_date+td(days=1))
                new_dataframe.loc[:,'CURTAILMENT END DATE TIME'] = new_dataframe.loc[:,'CURTAILMENT END DATE TIME'].apply(lambda t: min(t,effective_date))
                df = df.append(new_dataframe,ignore_index=True)
            else:
                pass
        return df
    
    def clear_all_downloads(self):
        '''
        Deletes all previously downloaded reports from the save location
        specified in the configuration file, and clears all data from the
        download log.
        '''
        download_dir = Path(d.today().strftime(self.download_path_template)).parent
        for p in download_dir.iterdir():
            p.unlink()
        self.logger.clear_log()
        self.logger.commit()